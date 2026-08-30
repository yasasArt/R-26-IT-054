from __future__ import annotations

from datetime import timedelta
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.database import connect, transaction
from app.time_utils import format_utc, parse_utc


def add_sample_data(client: TestClient, session: dict) -> None:
    started = parse_utc(session["started_at"])

    for elapsed in (15, 35):
        assert client.post(
            f"/api/sessions/{session['id']}/pieces",
            json={"completed_at": format_utc(started + timedelta(seconds=elapsed)), "event_source": "VALIDATION"},
        ).status_code == 201

    for event_type, elapsed in (("REWORK", 40), ("RESET", 70), ("DOWNTIME", 80), ("RESET", 125)):
        assert client.post(
            "/api/iot-events",
            json={
                "session_id": session["id"], "event_type": event_type,
                "event_source": "VALIDATION", "occurred_at": format_utc(started + timedelta(seconds=elapsed)),
            },
        ).status_code == 201


def test_combined_analytics_aggregates_sessions_cycles_and_iot_events(
    configured_client: TestClient, validation_session: dict
) -> None:
    add_sample_data(configured_client, validation_session)
    payload = configured_client.get("/api/analytics").json()

    assert payload["summary"]["session_count"] == 1
    assert payload["summary"]["total_pieces"] == 2
    assert payload["summary"]["average_cycle_seconds"] == 17.5
    assert payload["summary"]["rework_count"] == 1
    assert payload["summary"]["rework_seconds"] == 30.0
    assert payload["summary"]["downtime_count"] == 1
    assert payload["summary"]["downtime_seconds"] == 45.0
    assert len(payload["piece_events"]) == 2
    assert len(payload["iot_events"]) == 4


def test_analytics_supports_employee_session_line_and_mode_filters(
    configured_client: TestClient, employee: dict, validation_session: dict
) -> None:
    add_sample_data(configured_client, validation_session)
    matching = configured_client.get(
        f"/api/analytics?employee_id={employee['id']}&session_id={validation_session['id']}"
        "&sewing_line=Line%20A&session_mode=VALIDATION"
    ).json()
    assert matching["summary"]["total_pieces"] == 2

    other_line = configured_client.get("/api/analytics?sewing_line=Line%20B").json()
    assert other_line["summary"]["session_count"] == 0


def test_excel_export_contains_professional_industry_report_tabs(
    configured_client: TestClient, validation_session: dict
) -> None:
    add_sample_data(configured_client, validation_session)
    response = configured_client.get(f"/api/analytics/export.xlsx?session_id={validation_session['id']}")

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == [
        "Executive Summary", "Session Register", "Garment Cycle Detail",
        "IoT Event Register", "Employee Performance",
    ]
    assert workbook["Garment Cycle Detail"]["C5"].value == 15
    assert workbook["Session Register"]["E5"].value == "VALIDATION"
    assert workbook["IoT Event Register"]["C5"].value == "REWORK"
    assert workbook["Employee Performance"]["B5"].value == "Kavindi Perera"
    assert workbook["Executive Summary"]["F5"].value == validation_session["session_code"]
    assert workbook["Executive Summary"]["D13"].number_format == "0.0%"
    assert workbook["Session Register"]["I5"].value == "=IFERROR(H5/G5,0)"
    assert workbook["Session Register"]["I5"].number_format == "0.0%"
    assert workbook["Session Register"]["J5"].value == "=MAX(G5-H5,0)"
    assert workbook["Session Register"]["A4"].fill.fgColor.rgb == "00E9EDF2"
    assert len(workbook["Session Register"].conditional_formatting) == 0
    assert workbook["Session Register"]["T5"].is_date
    assert workbook["Garment Cycle Detail"]["A5"].value == validation_session["session_code"]
    assert workbook["Garment Cycle Detail"]["G5"].value == "Kavindi Perera"
    assert workbook["IoT Event Register"]["I5"].value == "Rework started"
    assert workbook["Employee Performance"]["I5"].value == 17.5
    assert workbook["Employee Performance"]["G5"].value == "=IFERROR(F5/E5,0)"


def test_empty_excel_export_remains_readable_and_valid(configured_client: TestClient) -> None:
    response = configured_client.get("/api/analytics/export.xlsx")
    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.content))
    assert workbook["Session Register"]["A5"].value == "No records match the selected filters."


def test_session_history_deletion_requires_exact_explicit_confirmation(
    configured_client: TestClient, validation_session: dict
) -> None:
    response = configured_client.post(
        "/api/sessions/delete-history", json={"confirmation": "delete everything"}
    )

    assert response.status_code == 422
    assert configured_client.get("/api/sessions").json()[0]["id"] == validation_session["id"]


def test_session_history_cannot_be_deleted_while_a_session_is_active(
    configured_client: TestClient, validation_session: dict
) -> None:
    add_sample_data(configured_client, validation_session)
    response = configured_client.post(
        "/api/sessions/delete-history", json={"confirmation": "DELETE SESSION DATA"}
    )

    assert response.status_code == 409
    assert "End the active session" in response.json()["detail"]
    analytics = configured_client.get("/api/analytics").json()
    assert analytics["summary"]["session_count"] == 1
    assert len(analytics["piece_events"]) == 2
    assert len(analytics["iot_events"]) == 4


def test_session_history_deletion_preserves_employees_devices_and_standalone_events(
    configured_client: TestClient, employee: dict, validation_session: dict
) -> None:
    add_sample_data(configured_client, validation_session)
    assert configured_client.post(
        f"/api/sessions/{validation_session['id']}/complete"
    ).status_code == 200
    standalone = configured_client.post(
        "/api/iot-events", json={"event_type": "RESET", "event_source": "VALIDATION"}
    )
    assert standalone.status_code == 201

    response = configured_client.post(
        "/api/sessions/delete-history", json={"confirmation": "DELETE SESSION DATA"}
    )

    assert response.status_code == 200
    assert response.json()["deleted_sessions"] == 1
    assert response.json()["deleted_piece_events"] == 2
    assert response.json()["deleted_iot_events"] == 4
    assert configured_client.get("/api/sessions").json() == []
    assert configured_client.get("/api/analytics").json()["summary"]["session_count"] == 0
    assert configured_client.get("/api/employees").json()[0]["id"] == employee["id"]
    assert configured_client.get("/api/device-configuration").json()["camera_tested"] is True

    retained_events = configured_client.get("/api/iot-events").json()
    assert len(retained_events) == 1
    assert retained_events[0]["id"] == standalone.json()["id"]
    assert retained_events[0]["session_id"] is None


def test_session_history_deletion_removes_legacy_nested_foreign_key_records(
    configured_client: TestClient, validation_session: dict
) -> None:
    connection = connect(configured_client.app.state.settings.database_path)
    try:
        with transaction(connection):
            connection.execute(
                "CREATE TABLE legacy_session_notes ("
                "id INTEGER PRIMARY KEY, "
                "session_id INTEGER REFERENCES production_sessions(id), "
                "note TEXT NOT NULL)"
            )
            connection.execute(
                "CREATE TABLE legacy_note_tags ("
                "id INTEGER PRIMARY KEY, "
                "note_id INTEGER NOT NULL REFERENCES legacy_session_notes(id), "
                "label TEXT NOT NULL)"
            )
            connection.execute(
                "INSERT INTO legacy_session_notes(id, session_id, note) VALUES (1, ?, 'legacy')",
                (validation_session["id"],),
            )
            connection.execute(
                "INSERT INTO legacy_session_notes(id, session_id, note) VALUES (2, NULL, 'standalone')"
            )
            connection.execute(
                "INSERT INTO legacy_note_tags(id, note_id, label) VALUES (1, 1, 'session')"
            )
    finally:
        connection.close()

    assert configured_client.post(
        f"/api/sessions/{validation_session['id']}/complete"
    ).status_code == 200
    response = configured_client.post(
        "/api/sessions/delete-history", json={"confirmation": "DELETE SESSION DATA"}
    )
    assert response.status_code == 200, response.text

    connection = connect(configured_client.app.state.settings.database_path)
    try:
        assert connection.execute("SELECT COUNT(*) FROM legacy_note_tags").fetchone()[0] == 0
        retained = connection.execute(
            "SELECT note FROM legacy_session_notes ORDER BY id"
        ).fetchall()
        assert [row["note"] for row in retained] == ["standalone"]
    finally:
        connection.close()
