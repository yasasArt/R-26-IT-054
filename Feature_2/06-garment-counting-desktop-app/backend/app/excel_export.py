from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

INK = "263448"
MUTED = "68758A"
ACCENT = "52647B"
WHITE = "FFFFFF"
PAPER = "F7F8FA"
HEADER = "E9EDF2"
LINE = "E2E7ED"
DATE_FORMAT = "yyyy-mm-dd hh:mm"
DECIMAL_FORMAT = "#,##0.0"
PERCENT_FORMAT = "0.0%"


def excel_datetime(value: str | None) -> datetime | None:
    if not value:
        return None

    parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
    return parsed


def duration_minutes(started_at: str, ended_at: str | None, generated_at: str) -> float:
    started = excel_datetime(started_at)
    ended = excel_datetime(ended_at or generated_at)
    assert started is not None and ended is not None
    return round(max(0.0, (ended - started).total_seconds()) / 60, 2)


def minutes(seconds: float | int | None) -> float:
    return round(float(seconds or 0) / 60, 2)


def add_title(worksheet: Any, title: str, subtitle: str, width: int) -> None:
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    worksheet.row_dimensions[1].height = 30
    worksheet.row_dimensions[2].height = 22
    worksheet.row_dimensions[3].height = 11

    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Aptos", size=18, color=INK, bold=True)
    title_cell.alignment = Alignment(vertical="center")

    subtitle_cell = worksheet.cell(row=2, column=1, value=subtitle)
    subtitle_cell.font = Font(name="Aptos", size=10, color=MUTED)
    subtitle_cell.alignment = Alignment(vertical="center")

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A5"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.print_options.horizontalCentered = True
    worksheet.print_title_rows = "1:4"
    worksheet.oddHeader.center.text = "Garment Counter | Production Analytics"
    worksheet.oddFooter.left.text = "Confidential production record"
    worksheet.oddFooter.right.text = "Page &P of &N"


def add_table(worksheet: Any, headers: list[str], rows: list[list[Any]], table_name: str) -> None:
    header_row = 4

    for column_number, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=column_number, value=header)
        cell.font = Font(name="Aptos", size=9, color=INK, bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.border = Border(bottom=Side(style="thin", color=LINE))
        cell.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        worksheet.column_dimensions[cell.column_letter].width = min(max(len(header) + 4, 15), 26)

    worksheet.row_dimensions[header_row].height = 33

    for row_number, values in enumerate(rows, start=header_row + 1):
        for column_number, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_number, column=column_number, value=value)
            cell.font = Font(name="Aptos", size=9, color=INK)
            cell.fill = PatternFill("solid", fgColor=WHITE if row_number % 2 else PAPER)
            cell.border = Border(bottom=Side(style="hair", color=LINE))
            cell.alignment = Alignment(
                vertical="center",
                horizontal=(
                    "right"
                    if isinstance(value, (int, float))
                    or (isinstance(value, str) and value.startswith("="))
                    else "left"
                ),
                indent=1,
            )

            if isinstance(value, datetime):
                cell.number_format = DATE_FORMAT
            elif isinstance(value, float):
                cell.number_format = DECIMAL_FORMAT
            elif isinstance(value, int):
                cell.number_format = "#,##0"

        worksheet.row_dimensions[row_number].height = 22

    if rows:
        table = Table(
            displayName=table_name,
            ref=f"A{header_row}:{worksheet.cell(row=header_row + len(rows), column=len(headers)).coordinate}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    else:
        empty = worksheet.cell(row=header_row + 1, column=1, value="No records match the selected filters.")
        empty.font = Font(name="Aptos", size=9, color=MUTED, italic=True)


def report_scope(payload: dict[str, Any]) -> dict[str, str]:
    filters = payload["filters"]
    sessions = payload["sessions"]
    selected_employee = next(
        (session for session in sessions if session["employee_id"] == filters.get("employee_id")),
        None,
    )
    selected_session = next(
        (session for session in sessions if session["id"] == filters.get("session_id")),
        None,
    )
    active_filters = any(value is not None and value != "" for value in filters.values())

    return {
        "scope": "Filtered production records" if active_filters else "All available production records",
        "employee": selected_employee["employee_name"] if selected_employee else "All employees",
        "session": selected_session["session_code"] if selected_session else "All sessions",
        "line": filters.get("sewing_line") or "All sewing lines",
        "mode": filters.get("session_mode") or "Production and validation",
        "start": filters.get("start_date") or "Earliest available",
        "end": filters.get("end_date") or "Latest available",
    }


def add_overview(workbook: Workbook, payload: dict[str, Any]) -> None:
    worksheet = workbook.active
    worksheet.title = "Executive Summary"
    add_title(worksheet, "PRODUCTION PERFORMANCE REPORT", "Clear, practical garment-production results", 8)

    for column in ("A", "E"):
        worksheet.column_dimensions[column].width = 19
    for column in ("B", "C", "D", "F", "G", "H"):
        worksheet.column_dimensions[column].width = 13

    scope = report_scope(payload)
    metadata = [
        ("Report scope", scope["scope"], "Generated (UTC)", excel_datetime(payload["generated_at"])),
        ("Employee", scope["employee"], "Session", scope["session"]),
        ("Sewing line", scope["line"], "Session type", scope["mode"]),
        ("From date", scope["start"], "Through date", scope["end"]),
    ]

    for row_number, values in enumerate(metadata, start=4):
        worksheet.merge_cells(start_row=row_number, start_column=2, end_row=row_number, end_column=4)
        worksheet.merge_cells(start_row=row_number, start_column=6, end_row=row_number, end_column=8)

        for column_number, value in zip((1, 2, 5, 6), values, strict=True):
            cell = worksheet.cell(row=row_number, column=column_number, value=value)
            cell.font = Font(
                name="Aptos", size=9, color=MUTED if column_number in (1, 5) else INK,
                bold=column_number in (2, 6),
            )
            cell.alignment = Alignment(vertical="center")
            if isinstance(value, datetime):
                cell.number_format = DATE_FORMAT

        worksheet.row_dimensions[row_number].height = 22

    worksheet.merge_cells("A9:H9")
    section_heading = worksheet["A9"]
    section_heading.value = "KEY PRODUCTION MEASURES"
    section_heading.font = Font(name="Aptos", size=10, color=ACCENT, bold=True)
    section_heading.alignment = Alignment(vertical="center")
    section_heading.border = Border(bottom=Side(style="thin", color=LINE))
    worksheet.row_dimensions[9].height = 25

    summary = payload["summary"]
    average_cycle = summary["average_cycle_seconds"]
    production_rate = round(3600 / average_cycle, 1) if average_cycle else 0.0
    overview_metrics = [
        ("Production sessions", summary["session_count"], "Completed sessions", summary["completed_session_count"]),
        ("Employees", summary["employee_count"], "Target pieces", summary["target_pieces"]),
        ("Produced pieces", summary["total_pieces"], "Pieces remaining", max(0, summary["target_pieces"] - summary["total_pieces"])),
        ("Target achieved", summary["achievement_percent"] / 100, "Avg cycle (seconds)", average_cycle),
        ("Sewing rate (pcs/hour)", production_rate, "Measured garments", len(payload["piece_events"])),
        ("Rework incidents", summary["rework_count"], "Rework time (minutes)", minutes(summary["rework_seconds"])),
        ("Downtime incidents", summary["downtime_count"], "Downtime time (minutes)", minutes(summary["downtime_seconds"])),
        ("Connection interruptions", summary["disconnect_count"], "Connection loss (minutes)", minutes(summary["disconnected_seconds"])),
    ]

    for row_number, values in enumerate(overview_metrics, start=10):
        worksheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=3)
        worksheet.merge_cells(start_row=row_number, start_column=5, end_row=row_number, end_column=7)

        for column_number in range(1, 9):
            worksheet.cell(row=row_number, column=column_number).fill = PatternFill(
                "solid", fgColor=PAPER if row_number % 2 == 0 else WHITE
            )

        for column_number, value in zip((1, 4, 5, 8), values, strict=True):
            cell = worksheet.cell(row=row_number, column=column_number, value=value)
            is_value = column_number in (4, 8)
            cell.font = Font(
                name="Aptos", size=12 if is_value else 9,
                color=INK if is_value else MUTED, bold=is_value,
            )
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if column_number == 4 else "right" if is_value else "left",
                indent=1,
            )
            if isinstance(value, float):
                cell.number_format = DECIMAL_FORMAT

        worksheet.row_dimensions[row_number].height = 29

    worksheet["D13"].number_format = PERCENT_FORMAT
    worksheet.merge_cells("A19:H19")
    note = worksheet["A19"]
    note.value = "Sewing rate is calculated from measured garment-cycle time. All timestamps are shown in UTC."
    note.font = Font(name="Aptos", size=9, color=MUTED, italic=True)
    note.alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[19].height = 28
    worksheet.freeze_panes = "A10"
    worksheet.print_title_rows = "1:2"
    worksheet.page_setup.orientation = "portrait"


def event_impact(event_type: str) -> str:
    return {
        "REWORK": "Rework started",
        "DOWNTIME": "Production paused",
        "RESET": "Normal production resumed",
        "DISCONNECTED": "Controller connection lost",
        "RECONNECTED": "Controller connection restored",
    }.get(event_type, event_type)


def build_workbook(payload: dict[str, Any]) -> bytes:
    workbook = Workbook()
    workbook.properties.creator = "Garment Counter"
    workbook.properties.title = "Garment Production Performance Report"
    workbook.properties.subject = "Filtered workstation production, employee, cycle, and controller analytics"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    add_overview(workbook, payload)

    sessions_by_id = {int(session["id"]): session for session in payload["sessions"]}
    cycles_by_employee: dict[int, list[float]] = defaultdict(list)
    for piece in payload["piece_events"]:
        session = sessions_by_id.get(int(piece["session_id"]))
        if session is not None:
            cycles_by_employee[int(session["employee_id"])].append(float(piece["cycle_seconds"]))

    session_headers = [
        "Session Code", "Employee Code", "Employee Name", "Sewing Line", "Mode", "Status",
        "Target Pieces", "Produced Pieces", "Target Achieved", "Pieces Remaining", "Avg Cycle (sec)",
        "Sewing Rate (pcs/hour)", "Session Length (min)", "Rework Incidents", "Rework Time (min)",
        "Downtime Incidents", "Downtime Time (min)", "Connection Losses", "Connection Loss (min)",
        "Started At (UTC)", "Ended At (UTC)", "Workstation",
    ]
    session_rows = []
    for row_number, session in enumerate(payload["sessions"], start=5):
        session_rows.append([
            session["session_code"], session["employee_code"], session["employee_name"],
            session["sewing_line"], session["session_mode"], session["status"],
            session["target_pieces"], session["total_pieces"],
            f"=IFERROR(H{row_number}/G{row_number},0)", f"=MAX(G{row_number}-H{row_number},0)",
            session["average_cycle_seconds"], f"=IFERROR(IF(K{row_number}>0,3600/K{row_number},0),0)",
            duration_minutes(session["started_at"], session["ended_at"], payload["generated_at"]),
            session["rework_count"], minutes(session["rework_seconds"]),
            session["downtime_count"], minutes(session["downtime_seconds"]),
            session["disconnect_count"], minutes(session["disconnected_seconds"]),
            excel_datetime(session["started_at"]), excel_datetime(session["ended_at"]),
            session["workstation_id"],
        ])

    sessions_sheet = workbook.create_sheet("Session Register")
    add_title(sessions_sheet, "SESSION PRODUCTION REGISTER", "Output, target achievement, and lost time by session", len(session_headers))
    add_table(sessions_sheet, session_headers, session_rows, "SessionRegister")
    for row_number in range(5, 5 + len(session_rows)):
        sessions_sheet[f"I{row_number}"].number_format = PERCENT_FORMAT
        sessions_sheet[f"L{row_number}"].number_format = DECIMAL_FORMAT

    piece_headers = [
        "Session Code", "Piece Number", "Cycle Time (sec)", "Vs Session Avg (sec)",
        "Sewing Started (UTC)", "Completed At (UTC)", "Employee Name", "Sewing Line", "Mode",
    ]
    piece_rows = []
    last_piece_row = 4 + len(payload["piece_events"])
    for row_number, event in enumerate(payload["piece_events"], start=5):
        session = sessions_by_id[int(event["session_id"])]
        piece_rows.append([
            session["session_code"], event["piece_number"], event["cycle_seconds"],
            f"=IFERROR(C{row_number}-AVERAGEIF($A$5:$A${last_piece_row},A{row_number},$C$5:$C${last_piece_row}),0)",
            excel_datetime(event["sewing_started_at"]), excel_datetime(event["completed_at"]),
            session["employee_name"], session["sewing_line"], session["session_mode"],
        ])

    piece_sheet = workbook.create_sheet("Garment Cycle Detail")
    add_title(piece_sheet, "GARMENT CYCLE DETAIL", "Measured garment completion times, including the first piece", len(piece_headers))
    add_table(piece_sheet, piece_headers, piece_rows, "GarmentCycleDetail")
    for row_number in range(5, 5 + len(piece_rows)):
        piece_sheet[f"D{row_number}"].number_format = DECIMAL_FORMAT

    iot_headers = [
        "Session Code", "Employee Name", "Event", "Previous Status", "New Status",
        "Recorded At (UTC)", "Controller", "Session Mode", "Production Impact",
    ]
    iot_rows = []
    for event in payload["iot_events"]:
        session = sessions_by_id[int(event["session_id"])]
        iot_rows.append([
            session["session_code"], session["employee_name"], event["event_type"],
            event["mode_before"], event["mode_after"], excel_datetime(event["occurred_at"]),
            event["device_name"] or "Operator controller", session["session_mode"],
            event_impact(event["event_type"]),
        ])

    iot_sheet = workbook.create_sheet("IoT Event Register")
    add_title(iot_sheet, "OPERATOR CONTROLLER ACTIVITY", "Rework, downtime, reset, and connection history", len(iot_headers))
    add_table(iot_sheet, iot_headers, iot_rows, "IoTEventRegister")

    employee_headers = [
        "Employee Code", "Employee Name", "Sewing Line", "Sessions", "Target Pieces",
        "Produced Pieces", "Target Achieved", "Pieces Remaining", "Avg Cycle (sec)",
        "Sewing Rate (pcs/hour)", "Rework Incidents", "Rework Time (min)",
        "Downtime Incidents", "Downtime Time (min)",
    ]
    employee_rows = []
    for row_number, employee in enumerate(payload["employees"], start=5):
        cycle_values = cycles_by_employee.get(int(employee["employee_id"]), [])
        average_cycle = round(sum(cycle_values) / len(cycle_values), 2) if cycle_values else None
        employee_rows.append([
            employee["employee_code"], employee["employee_name"], employee["sewing_line"],
            employee["session_count"], employee["target_pieces"], employee["total_pieces"],
            f"=IFERROR(F{row_number}/E{row_number},0)", f"=MAX(E{row_number}-F{row_number},0)",
            average_cycle, f"=IFERROR(IF(I{row_number}>0,3600/I{row_number},0),0)",
            employee["rework_count"], minutes(employee["rework_seconds"]),
            employee["downtime_count"], minutes(employee["downtime_seconds"]),
        ])

    employee_sheet = workbook.create_sheet("Employee Performance")
    add_title(employee_sheet, "EMPLOYEE PERFORMANCE", "Target delivery, measured sewing rate, and production interruptions", len(employee_headers))
    add_table(employee_sheet, employee_headers, employee_rows, "EmployeePerformance")
    for row_number in range(5, 5 + len(employee_rows)):
        employee_sheet[f"G{row_number}"].number_format = PERCENT_FORMAT
        employee_sheet[f"J{row_number}"].number_format = DECIMAL_FORMAT

    for worksheet in workbook.worksheets[1:]:
        worksheet.page_setup.orientation = "landscape"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
