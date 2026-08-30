from datetime import UTC, datetime # type: ignore
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.analytics import AnalyticsResponse

NAVY = "1F2937"
TEAL = "0F766E"
LIGHT_TEAL = "CCFBF1"
LIGHT_GREY = "F3F4F6"
MID_GREY = "D1D5DB"
WHITE = "FFFFFF"
TEXT = "111827"
THIN_GREY = Side(style="thin", color=MID_GREY)


def _excel_datetime(value: datetime | None) -> datetime | None:
    """Excel stores timezone-free values; exports consistently use UTC."""

    if value is None:
        return None
    return value.astimezone(UTC).replace(tzinfo=None)


class ExcelService:
    def __init__(self, analytics: AnalyticsResponse) -> None:
        self.analytics = analytics

    @staticmethod
    def _title(sheet: Worksheet, title: str, last_column: int) -> None:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
        cell = sheet.cell(1, 1, title)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True, size=16)
        cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 28

    @staticmethod
    def _header(sheet: Worksheet, row: int, columns: int) -> None:
        for cell in sheet.iter_cols(
            min_row=row,
            max_row=row,
            min_col=1,
            max_col=columns,
        ):
            current = cell[0]
            current.fill = PatternFill("solid", fgColor=NAVY)
            current.font = Font(color=WHITE, bold=True)
            current.alignment = Alignment(horizontal="center", vertical="center")
            current.border = Border(bottom=THIN_GREY)
        sheet.row_dimensions[row].height = 24

    @staticmethod
    def _stripe(sheet: Worksheet, start_row: int, end_row: int, columns: int) -> None:
        for row in range(start_row, end_row + 1):
            fill = PatternFill(
                "solid",
                fgColor=LIGHT_GREY if row % 2 == 0 else WHITE,
            )
            for cell in sheet[row][:columns]:
                cell.fill = fill
                cell.font = Font(color=TEXT)
                cell.border = Border(bottom=THIN_GREY)
                cell.alignment = Alignment(vertical="top")

    @staticmethod
    def _total_row(sheet: Worksheet, row: int, columns: int) -> None:
        for cell in sheet[row][:columns]:
            cell.fill = PatternFill("solid", fgColor=LIGHT_TEAL)
            cell.font = Font(color=TEXT, bold=True)
            cell.border = Border(top=Side(style="medium", color=TEAL))

    @staticmethod
    def _set_widths(sheet: Worksheet, widths: list[float]) -> None:
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    @staticmethod
    def _base_sheet(sheet: Worksheet) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.pageSetUpPr.fitToPage = True # type: ignore
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.outlinePr.summaryBelow = True # type: ignore

    def _management_summary(self, workbook: Workbook) -> None:
        sheet = workbook.create_sheet("Management Summary")
        self._base_sheet(sheet)
        self._title(sheet, "Garment Production Management Summary", 6)
        sheet.freeze_panes = "A4"

        sheet["A3"] = "Generated at (UTC)"
        sheet["B3"] = _excel_datetime(self.analytics.generated_at)
        sheet["B3"].number_format = "yyyy-mm-dd hh:mm:ss"
        sheet["D3"] = "Applied filters"
        active_filters = [
            f"{name}={value.value if hasattr(value, 'value') else value}"
            for name, value in self.analytics.filters.model_dump().items()
            if value is not None
        ]
        sheet["E3"] = ", ".join(active_filters) if active_filters else "All sessions"
        sheet.merge_cells("E3:F3")

        summary = self.analytics.summary
        cards = [
            ("Sessions", summary.total_sessions, "Employees", summary.unique_employees),
            ("Target pieces", summary.target_pieces, "Confirmed pieces", summary.confirmed_pieces),
            ("Remaining pieces", summary.remaining_pieces, "Achievement", summary.achievement_percent / 100),
            ("Average cycle", summary.average_cycle_seconds, "Completed sessions", summary.completed_sessions),
            ("Rework count", summary.rework_count, "Rework duration", summary.rework_duration_seconds),
            ("Downtime count", summary.downtime_count, "Downtime duration", summary.downtime_duration_seconds),
        ]
        start_row = 5
        for offset, (left_label, left_value, right_label, right_value) in enumerate(cards):
            row = start_row + offset
            sheet.cell(row, 1, left_label)
            sheet.cell(row, 2, left_value)
            sheet.cell(row, 4, right_label)
            sheet.cell(row, 5, right_value)
            for column in (1, 4):
                sheet.cell(row, column).font = Font(bold=True, color=TEXT)
                sheet.cell(row, column).fill = PatternFill("solid", fgColor=LIGHT_GREY)
            for column in (2, 5):
                sheet.cell(row, column).font = Font(bold=True, color=TEAL, size=12)
                sheet.cell(row, column).fill = PatternFill("solid", fgColor=WHITE)
                sheet.cell(row, column).alignment = Alignment(horizontal="right")

        sheet["E7"].number_format = "0.0%"
        sheet["B8"].number_format = '0.000 "s"'
        sheet["E9"].number_format = '0.000 "s"'
        sheet["E10"].number_format = '0.000 "s"'

        table_row = 13
        headers = ["Status", "Sessions", "Share"]
        sheet.cell(table_row, 1, headers[0])
        sheet.cell(table_row, 2, headers[1])
        sheet.cell(table_row, 3, headers[2])
        rows = [
            ("Active", summary.active_sessions),
            ("Completed", summary.completed_sessions),
            ("Cancelled", summary.cancelled_sessions),
        ]
        for index, (label, value) in enumerate(rows, start=table_row + 1):
            sheet.cell(index, 1, label)
            sheet.cell(index, 2, value)
            sheet.cell(index, 3, f"=IF($B$5=0,0,B{index}/$B$5)")
            sheet.cell(index, 3).number_format = "0.0%"
        self._header(sheet, table_row, 3)
        self._stripe(sheet, table_row + 1, table_row + len(rows), 3)
        sheet.auto_filter.ref = f"A{table_row}:C{table_row + len(rows)}"
        self._set_widths(sheet, [24, 18, 16, 24, 28, 18])
        sheet.print_title_rows = "1:3"

    def _session_register(self, workbook: Workbook) -> None:
        sheet = workbook.create_sheet("Session Register")
        self._base_sheet(sheet)
        headers = [
            "Session ID",
            "Employee No.",
            "Employee",
            "Sewing Line",
            "Mode",
            "Status",
            "Started (UTC)",
            "Ended (UTC)",
            "Target Pieces",
            "Confirmed Pieces",
            "Remaining Pieces",
            "Achievement",
            "Avg Cycle (s)",
            "Rework Count",
            "Rework Duration (s)",
            "Downtime Count",
            "Downtime Duration (s)",
        ]
        self._title(sheet, "Production Session Register", len(headers))
        header_row = 3
        for column, value in enumerate(headers, start=1):
            sheet.cell(header_row, column, value)
        self._header(sheet, header_row, len(headers))

        first_data = header_row + 1
        for row, report in enumerate(self.analytics.sessions, start=first_data):
            values = [
                report.session_id,
                report.employee_number,
                report.employee_name,
                report.sewing_line,
                report.session_mode.value,
                report.session_status.value,
                _excel_datetime(report.started_at),
                _excel_datetime(report.ended_at),
                report.target_pieces,
                report.confirmed_pieces,
                report.remaining_pieces,
                report.achievement_percent / 100,
                report.average_cycle_seconds,
                report.rework.count,
                report.rework.duration_seconds,
                report.downtime.count,
                report.downtime.duration_seconds,
            ]
            for column, value in enumerate(values, start=1):
                sheet.cell(row, column, value)

        last_data = first_data + len(self.analytics.sessions) - 1
        if self.analytics.sessions:
            self._stripe(sheet, first_data, last_data, len(headers))
            sheet.auto_filter.ref = f"A{header_row}:Q{last_data}"
            total_row = last_data + 2
            sheet.cell(total_row, 1, "Totals / Average")
            for column in (9, 10, 11, 14, 15, 16, 17):
                letter = get_column_letter(column)
                sheet.cell(total_row, column, f"=SUM({letter}{first_data}:{letter}{last_data})")
            sheet.cell(total_row, 12, f"=IF(I{total_row}=0,0,J{total_row}/I{total_row})")
            sheet.cell(
                total_row,
                13,
                f"=IF(J{total_row}=0,0,"
                f"SUMPRODUCT(J{first_data}:J{last_data},M{first_data}:M{last_data})/"
                f"J{total_row})",
            )
            self._total_row(sheet, total_row, len(headers))
            sheet.cell(total_row, 12).number_format = "0.0%"
            sheet.cell(total_row, 13).number_format = '0.000 "s"'
        else:
            sheet.auto_filter.ref = f"A{header_row}:Q{header_row}"

        for row in range(first_data, max(first_data, last_data) + 1):
            for column in (7, 8):
                sheet.cell(row, column).number_format = "yyyy-mm-dd hh:mm:ss"
            sheet.cell(row, 12).number_format = "0.0%"
            for column in (13, 15, 17):
                sheet.cell(row, column).number_format = '0.000 "s"'
        sheet.freeze_panes = "A4"
        sheet.print_title_rows = "1:3"
        self._set_widths(
            sheet,
            [11, 15, 24, 15, 13, 13, 21, 21, 14, 17, 16, 14, 15, 13, 19, 16, 21],
        )

    def _piece_details(self, workbook: Workbook) -> None:
        sheet = workbook.create_sheet("Piece Details")
        self._base_sheet(sheet)
        headers = [
            "Session ID",
            "Employee",
            "Piece Number",
            "Cycle (s)",
            "Confidence",
            "Source",
            "Sewing Started (UTC)",
            "Completed (UTC)",
        ]
        self._title(sheet, "Confirmed Garment Piece Details", len(headers))
        header_row = 3
        for column, value in enumerate(headers, start=1):
            sheet.cell(header_row, column, value)
        self._header(sheet, header_row, len(headers))

        row = header_row + 1
        for report in self.analytics.sessions:
            for piece in report.individual_cycle_times:
                values = [
                    report.session_id,
                    report.employee_name,
                    piece.piece_number,
                    piece.cycle_seconds,
                    piece.confidence,
                    piece.event_source.value,
                    _excel_datetime(piece.sewing_started_at),
                    _excel_datetime(piece.completed_at),
                ]
                for column, value in enumerate(values, start=1):
                    sheet.cell(row, column, value)
                row += 1

        first_data = header_row + 1
        last_data = row - 1
        if last_data >= first_data:
            self._stripe(sheet, first_data, last_data, len(headers))
            sheet.auto_filter.ref = f"A{header_row}:H{last_data}"
            total_row = last_data + 2
            sheet.cell(total_row, 1, "Count / Average")
            sheet.cell(total_row, 3, f"=COUNT(C{first_data}:C{last_data})")
            sheet.cell(total_row, 4, f"=AVERAGE(D{first_data}:D{last_data})")
            self._total_row(sheet, total_row, len(headers))
            sheet.cell(total_row, 4).number_format = '0.000 "s"'
        else:
            sheet.auto_filter.ref = f"A{header_row}:H{header_row}"

        for data_row in range(first_data, last_data + 1):
            sheet.cell(data_row, 4).number_format = '0.000 "s"'
            sheet.cell(data_row, 5).number_format = "0.0%"
            for column in (7, 8):
                sheet.cell(data_row, column).number_format = "yyyy-mm-dd hh:mm:ss"
        sheet.freeze_panes = "A4"
        sheet.print_title_rows = "1:3"
        self._set_widths(sheet, [12, 24, 14, 14, 13, 18, 23, 23])

    def _rework_downtime(self, workbook: Workbook) -> None:
        sheet = workbook.create_sheet("Rework and Downtime")
        self._base_sheet(sheet)
        headers = [
            "Session ID",
            "Employee",
            "Event ID",
            "Event Type",
            "Mode Before",
            "Mode After",
            "Device",
            "Source",
            "Occurred (UTC)",
            "Closed Duration (s)",
        ]
        self._title(sheet, "Rework and Downtime Event Register", len(headers))
        header_row = 3
        for column, value in enumerate(headers, start=1):
            sheet.cell(header_row, column, value)
        self._header(sheet, header_row, len(headers))

        row = header_row + 1
        for report in self.analytics.sessions:
            for event in report.operator_events:
                values = [
                    report.session_id,
                    report.employee_name,
                    event.event_id,
                    event.event_type.value,
                    event.mode_before.value,
                    event.mode_after.value,
                    event.device_name,
                    event.event_source.value,
                    _excel_datetime(event.occurred_at),
                    event.closed_mode_duration_seconds,
                ]
                for column, value in enumerate(values, start=1):
                    sheet.cell(row, column, value)
                row += 1

        first_data = header_row + 1
        last_data = row - 1
        if last_data >= first_data:
            self._stripe(sheet, first_data, last_data, len(headers))
            sheet.auto_filter.ref = f"A{header_row}:J{last_data}"
            total_row = last_data + 2
            sheet.cell(total_row, 1, "Closed duration total")
            sheet.cell(total_row, 10, f"=SUM(J{first_data}:J{last_data})")
            self._total_row(sheet, total_row, len(headers))
            sheet.cell(total_row, 10).number_format = '0.000 "s"'
        else:
            sheet.auto_filter.ref = f"A{header_row}:J{header_row}"

        for data_row in range(first_data, last_data + 1):
            sheet.cell(data_row, 9).number_format = "yyyy-mm-dd hh:mm:ss"
            sheet.cell(data_row, 10).number_format = '0.000 "s"'
        sheet.freeze_panes = "A4"
        sheet.print_title_rows = "1:3"
        self._set_widths(sheet, [12, 24, 12, 15, 15, 15, 22, 23, 22, 21])

    def build(self) -> bytes:
        workbook = Workbook()
        workbook.remove(workbook.active) # type: ignore
        workbook.properties.title = "Garment Counter Analytics"
        workbook.properties.subject = "Production session analytics export"
        workbook.properties.creator = "Garment Counter Backend"

        self._management_summary(workbook)
        self._session_register(workbook)
        self._piece_details(workbook)
        self._rework_downtime(workbook)

        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
